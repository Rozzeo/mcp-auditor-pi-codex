"""Connector approval matrix — one row per callable operation, for InfoSec review.

An audit answers "is this server dangerous?". A review board asks a different
question: enumerate every operation this connector exposes, say what kind of
access each one is, and let a human approve them individually. This module
produces that table.

Columns A-D (connector, action, type, description) are generated and should not
be hand-edited — regenerate instead. E and F (recommendation, InfoSec status)
are seeded with a defensible default and are meant to be overridden by people;
G (comments) is left for the reviewer.

Nothing here runs the connector. Operations come from what the extractor already
parses: a saved `tools/list` response, a JSON manifest, or a server's source.
"""

from __future__ import annotations

import csv
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable

import yaml

from .capabilities import DESTRUCTIVE_CAPABILITIES, MUTATING_CAPABILITIES
from .types import Tool

# The fixed core vocabulary: three labels, because that is the question a
# review board actually answers — does this operation observe state, change it,
# or destroy it. Delete is the destructive form of write and is split out only
# because it is the one that cannot be undone by writing again.
#
# Finer distinctions a particular connector cares about ("Scheduling",
# "Mailbox settings", "Search", "Discovery") go in an overrides file rather
# than into this list, so matrices stay comparable between connectors.
READ, WRITE, DELETE = "Read", "Write", "Delete"
CORE_TYPES = (READ, WRITE, DELETE)

# Verb families as WHOLE WORDS. Names arrive in three shapes -- `posts.list`,
# `outlook_send_mail`, `wpcom-domain-update-dns-records` -- so classification
# tokenizes the name and matches tokens exactly.
#
# Exact tokens, never prefixes: `^share` would classify every `sharepoint_*`
# tool as a send, and `^mail` would do the same to `set-mail-service`. For the
# same reason "email"/"mail" are absent — in real tool names they are nouns
# (`outlook_email_search` reads), and the verb form is already covered by "send".
#
# Sending is a write: it leaves the system changed and observable by someone
# else. Searching and describing are reads: they return state without altering it.
_VERB_TOKENS: tuple[tuple[str, frozenset], ...] = (
    (DELETE, frozenset({"delete", "remove", "destroy", "drop", "purge", "trash",
                        "erase", "revoke", "uninstall", "unpublish"})),
    (WRITE, frozenset({"send", "notify", "dispatch", "forward", "publish",
                       "broadcast", "invite",
                       "create", "update", "add", "set", "modify", "edit", "rename",
                       "move", "copy", "upload", "patch", "put", "write", "change",
                       "activate", "deactivate", "launch", "restore", "purchase",
                       "register", "renew", "install", "enable",
                       "disable", "assign", "approve", "submit", "import", "sync",
                       "refresh", "reset", "untrash", "respond"})),
    (READ, frozenset({"search", "find", "query", "lookup",
                      "describe", "discover", "catalog", "schema", "capabilities",
                      "list", "get", "read", "fetch", "view", "show", "status",
                      "preview", "download", "export", "check", "count", "resolve"})),
)


@dataclass
class ActionRow:
    """One callable operation of one connector."""

    connector: str
    action: str
    type: str
    description: str
    recommends: str = ""
    status: str = ""
    comments: str = ""
    platform: str = ""
    # Not exported: how the type was decided, for --explain and for tests.
    reason: str = ""

    def cells(self, with_platform: bool) -> list[str]:
        row = [self.connector, self.action, self.type, self.description,
               self.recommends, self.status, self.comments]
        if with_platform:
            row.append(self.platform)
        return row


HEADER = ["connector", "action", "type", "description",
          "AI team recomends", "Status - Approved by InfoSec", "Comments"]


@dataclass
class Overrides:
    """Connector-specific labels, so a domain vocabulary needs no code change."""

    exact: dict[str, str] = field(default_factory=dict)
    patterns: list[tuple[re.Pattern[str], str]] = field(default_factory=list)

    @classmethod
    def load(cls, path: str | Path | None) -> "Overrides":
        if not path:
            return cls()
        data = yaml.safe_load(Path(path).read_text(encoding="utf-8")) or {}
        exact = {str(k): str(v) for k, v in (data.get("types") or {}).items()}
        pats = []
        for entry in data.get("patterns") or []:
            if isinstance(entry, dict) and entry.get("match") and entry.get("type"):
                pats.append((re.compile(str(entry["match"]), re.IGNORECASE), str(entry["type"])))
        return cls(exact=exact, patterns=pats)

    def lookup(self, action: str) -> str | None:
        if action in self.exact:
            return self.exact[action]
        operation = _operation_of(action)
        if operation in self.exact:
            return self.exact[operation]
        for rx, label in self.patterns:
            if rx.search(action):
                return label
        return None


def _operation_of(action: str) -> str:
    """The part that carries the verb: `facade -> posts.list` becomes `posts.list`.

    Also tolerates the hand-written `-> action: list / describe` shape that
    appears in matrices built by reading vendor docs.
    """
    tail = action.split("->")[-1].strip()
    tail = re.sub(r"^(?:action|operation|method)\s*:\s*", "", tail, flags=re.IGNORECASE)
    return tail.split("/")[0].strip() if "/" in tail else tail


def _verb_of(operation: str) -> str:
    """Last dotted segment, so `posts.list` tests as `list`, not `posts`."""
    return operation.rsplit(".", 1)[-1].strip()


def _tokens_of(operation: str) -> list[str]:
    """Every word in the name, most-significant last.

    Names come in three shapes — `posts.list`, `outlook_send_mail`, and
    `wpcom-domain-update-dns-records`. Anchoring a verb at the start of the
    whole string only works for the first, so each token is tested separately;
    the trailing tokens are checked first because the verb usually sits at the
    end of a hyphenated vendor name.
    """
    return [t for t in re.split(r"[^A-Za-z0-9]+", operation) if t]


def classify(tool_or_action, description: str = "", annotations: dict | None = None,
             capabilities: Iterable | None = None, overrides: Overrides | None = None) -> tuple[str, str]:
    """Return (type, reason). Deterministic and order-dependent — see _VERB_RULES.

    Precedence: an explicit override, then a declared destructive annotation,
    then the verb in the operation name, then a declared read-only annotation,
    then statically inferred capabilities. Names beat inferred capabilities
    because a doc-derived tool list has no body to infer from at all, and the
    verb is what a reviewer reads anyway.
    """
    if isinstance(tool_or_action, Tool):
        action = tool_or_action.name
        description = description or tool_or_action.description
        annotations = tool_or_action.annotations if annotations is None else annotations
        capabilities = tool_or_action.capabilities if capabilities is None else capabilities
    else:
        action = str(tool_or_action)
    annotations = annotations or {}
    caps = {c.capability for c in (capabilities or [])}

    if overrides:
        hit = overrides.lookup(action)
        if hit:
            return hit, "overrides file"

    operation = _operation_of(action)

    if annotations.get("destructiveHint") is True:
        return DELETE, "destructiveHint: true"

    # Tokens are scanned EARLIEST first because vendor names read
    # <namespace>-<verb>-<object>: `set-mail-service` is a write on the mail
    # service, not a send of mail.
    for token in _tokens_of(operation):
        low_token = token.lower()
        for label, verbs in _VERB_TOKENS:
            if low_token in verbs:
                return label, f"token '{low_token}'"

    if annotations.get("readOnlyHint") is True:
        return READ, "readOnlyHint: true"
    if caps & DESTRUCTIVE_CAPABILITIES:
        return DELETE, "destructive capability in body"
    if caps & MUTATING_CAPABILITIES:
        return WRITE, "mutating capability in body"
    if caps:
        return READ, "read-only capabilities in body"
    return READ, "no signal — defaulted, verify manually"


# Default recommendation per type. The reviewer overrides column E; this is a
# starting position, not a verdict. Delete defaults to refusal because it is
# the operation that cannot be undone by issuing another one.
_RECOMMENDS = {
    READ: "Approved",
    WRITE: "Approved (with confirmation)",
    DELETE: "No",
}


def _recommend(type_label: str) -> str:
    if type_label in _RECOMMENDS:
        return _RECOMMENDS[type_label]
    low = type_label.lower()
    if "destructive" in low or "delete" in low:
        return "No"
    if any(w in low for w in ("write", "send", "create", "update")):
        return "Approved (with confirmation)"
    return "Approved"


def expand_facade(tool: Tool) -> list[str]:
    """Facade tools expose many operations behind one MCP tool. Return the
    `tool -> resource.action` names when the schema enumerates them, else [].

    Only literal enums are expanded: guessing operations a schema does not list
    would put rows in front of a review board that no one can verify.
    """
    schema = tool.schema if isinstance(tool.schema, dict) else {}
    props = schema.get("properties")
    if not isinstance(props, dict):
        return []
    for key in ("operation", "action", "method", "command", "resource_action"):
        spec = props.get(key)
        if isinstance(spec, dict) and isinstance(spec.get("enum"), list):
            values = [str(v) for v in spec["enum"] if isinstance(v, (str, int))]
            if values:
                return [f"{tool.name} -> {v}" for v in values]
    return []


def build_matrix(tools: Iterable[Tool], connector: str, *, status: str = "Pending InfoSec review",
                 platform: str = "", overrides: Overrides | None = None,
                 prefill: bool = True) -> list[ActionRow]:
    """Turn extracted tools into review rows, expanding facade tools in place."""
    rows: list[ActionRow] = []
    for tool in tools:
        actions = expand_facade(tool) or [tool.name]
        for action in actions:
            label, reason = classify(action, tool.description, tool.annotations,
                                     tool.capabilities, overrides)
            # The recommendation follows the CORE type, not the override label.
            # A connector-specific name like "Mailbox settings" says nothing
            # about risk, and letting it drive the default would silently
            # downgrade a write to "Approved".
            core, _ = classify(action, tool.description, tool.annotations, tool.capabilities)
            rows.append(ActionRow(
                connector=connector,
                action=action,
                type=label,
                description=tool.description.replace("\n", " ").strip(),
                recommends=_recommend(core if label not in CORE_TYPES else label) if prefill else "",
                status=status if prefill else "",
                platform=platform,
                reason=reason,
            ))
    return rows


def summarize(rows: Iterable[ActionRow]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for row in rows:
        counts[row.type] = counts.get(row.type, 0) + 1
    return dict(sorted(counts.items(), key=lambda kv: -kv[1]))


def write_csv(rows: list[ActionRow], path: str | Path, platform: bool = False) -> None:
    header = HEADER + (["platform"] if platform else [])
    with open(path, "w", encoding="utf-8-sig", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(header)
        for row in rows:
            w.writerow(row.cells(platform))


def write_xlsx(rows: list[ActionRow], path: str | Path, platform: bool = False) -> None:
    """One flat sheet, matching the shape review boards already use.

    Requires openpyxl; write_csv is the dependency-free fallback.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - depends on optional extra
        raise RuntimeError(
            "Writing .xlsx needs openpyxl. Install it with: pip install 'mcp-auditor[xlsx]' "
            "(or use --format csv)."
        ) from exc

    header = HEADER + (["platform"] if platform else [])
    wb = Workbook()
    ws = wb.active
    ws.title = "Connectors"
    ws.append(header)

    head_fill = PatternFill("solid", fgColor="4472C4")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = head_fill
    ws.freeze_panes = "A2"

    # Same colour language as the sheets this replaces: write is amber,
    # destructive is red, so the risky rows are findable without reading.
    amber = PatternFill("solid", fgColor="FFE699")
    red = PatternFill("solid", fgColor="F4B0B0")
    for row in rows:
        ws.append(row.cells(platform))
        low = row.type.lower()
        fill = red if ("delete" in low or "destructive" in low) else (
            amber if low in ("write", "send") or "write" in low else None)
        if fill:
            ws.cell(row=ws.max_row, column=3).fill = fill

    widths = [22, 46, 16, 62, 26, 28, 34, 30]
    for i, width in enumerate(widths[:len(header)], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(path)
